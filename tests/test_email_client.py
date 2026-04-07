from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
import pytest

from core.email_client import AsyncEmailClient


class TestAsyncEmailClient:
    def test_build_request_splits_addresses(self):
        client = AsyncEmailClient()
        request = client.build_request(
            {
                "email_from_address": "sender@example.com",
                "email_from_auth_code": "secret",
                "email_to_addresses": "a@example.com, b@example.com",
                "email_cc_addresses": "cc1@example.com, cc2@example.com",
            },
            subject="test",
            plain_text_body="hello",
        )
        assert list(request.to_addresses) == ["a@example.com", "b@example.com"]
        assert list(request.cc_addresses) == ["cc1@example.com", "cc2@example.com"]

    def test_build_request_requires_sender(self):
        client = AsyncEmailClient()
        with pytest.raises(ValueError):
            client.build_request(
                {
                    "email_from_auth_code": "secret",
                    "email_to_addresses": "a@example.com",
                },
                subject="test",
                plain_text_body="hello",
            )

    def test_product_name_drives_summary_subject(self):
        client = AsyncEmailClient()
        request = client.build_request(
            {
                "site_title": "My Sentinel",
                "email_from_address": "sender@example.com",
                "email_from_auth_code": "secret",
                "email_to_addresses": "a@example.com",
            },
            subject=f"{client._product_name({'site_title': 'My Sentinel'})} 每日总结 2026-01-01",
            plain_text_body="hello",
        )
        assert request.subject == "My Sentinel 每日总结 2026-01-01"

    async def test_send_daily_summary_email_attaches_single_xlsx_table(self):
        client = AsyncEmailClient()
        captured = {}

        async def fake_send_email(request):
            captured["request"] = request
            return {"status": "SUCCESS"}

        client.send_email = fake_send_email
        await client.send_daily_summary_email(
            {
                "site_title": "My Sentinel",
                "email_from_address": "sender@example.com",
                "email_from_auth_code": "secret",
                "email_to_addresses": "a@example.com",
            },
            summary_text="hello",
            until_iso="2026-01-01T12:00:00+00:00",
            visits=[
                {
                    "source_name": "Cam1",
                    "plate": "ABC123",
                    "enter_time": "2026-01-01T00:00:00+00:00",
                    "exit_time": "2026-01-01T01:00:00+00:00",
                    "missing_actions": ["HandOverKeys"],
                },
                {
                    "source_name": "Cam2",
                    "plate": "",
                    "enter_time": "2026-01-01T02:00:00+00:00",
                    "exit_time": "2026-01-01T03:00:00+00:00",
                    "missing_actions": [],
                },
            ],
        )
        request = captured["request"]
        assert request.subject == "2026年01月01日AI货台分析报告-有异常"
        assert len(request.attachments) == 1
        attachment = request.attachments[0]
        assert attachment.filename == "truck-daily-summary-2026-01-01.xlsx"
        workbook = load_workbook(BytesIO(attachment.data))
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        assert rows == [
            ["序号", "区域", "回放位置或IP", "抽查内容/类型", "AI视觉分析结果"],
            [
                "1",
                None,
                "Cam1",
                "货台检查\n车牌号：ABC123\n到达时间：2026-01-01 00:00:00\n离开时间：2026-01-01 01:00:00",
                "未执行 上交钥匙",
            ],
            [
                "2",
                None,
                "Cam2",
                "货台检查\n车牌号：未识别\n到达时间：2026-01-01 02:00:00\n离开时间：2026-01-01 03:00:00",
                "无异常",
            ],
        ]
        assert "请查收附件" not in request.plain_text_body
        assert request.plain_text_body.startswith("序号\t区域\t回放位置或IP")
        assert "<table" in request.html_body
